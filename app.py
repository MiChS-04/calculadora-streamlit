import streamlit as st
import pandas as pd
import io
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. CONFIGURACIÓN DE LA PÁGINA
st.set_page_config(
    page_title="Calculadora Integral de Materiales - CAPECO",
    page_icon="🏗️",
    layout="wide"
)

st.title("🏗️ Sistema Integral de Presupuestos de Construcción (Normativa CAPECO)")
st.markdown("""
Esta plataforma realiza el análisis global de materiales para una edificación, integrando **muros, acabados, estructuras de concreto y acero estructural**, aplicando las mermas y rendimientos estándar de la Cámara Peruana de la Construcción (CAPECO).
""")
st.write("---")

# 2. ENTRADAS DE DATOS EN LA BARRA LATERAL (INPUTS)
st.sidebar.header("📐 Parámetros Globales del Proyecto")
num_pisos = st.sidebar.number_input("Número de Pisos de la Edificación:", min_value=1, value=1, step=1)
area_construccion = st.sidebar.number_input("Área por Piso / Techo (m²):", min_value=1.0, value=60.0, step=5.0)
area_pared = st.sidebar.number_input("Área de Pared a Enchapar Total (m²):", min_value=0.0, value=20.0, step=5.0)

# -----------------------------------------------------------------------------
# INTEGRACIÓN DE LA OPCIÓN A: CÁLCULO Y VISUALIZACIÓN EN LA BARRA LATERAL
# -----------------------------------------------------------------------------
area_total_construir = num_pisos * area_construccion
st.sidebar.info(f"**🏢 Área Total a Construir:** {area_total_construir:.2f} m²")

# MÓDULO DE LADRILLOS DE MURO (CON OPCIÓN PERSONALIZADA)
st.sidebar.write("---")
st.sidebar.subheader("🧱 Ladrillo para Muros")
tipo_ladrillo = st.sidebar.selectbox(
    "Formato de Ladrillo de Muro:",
    ["King Kong Estándar (24x13x9 cm)", "Ladrillo Pandereta (23x11x9 cm)", "Otro (Personalizado)"]
)

if tipo_ladrillo == "King Kong Estándar (24x13x9 cm)":
    L_lad, H_lad = 0.24, 0.09
elif tipo_ladrillo == "Ladrillo Pandereta (23x11x9 cm)":
    L_lad, H_lad = 0.23, 0.09
else:
    col_l1, col_l2 = st.sidebar.columns(2)
    l_cm = col_l1.number_input("Largo (cm):", min_value=1.0, value=24.0, step=0.5)
    h_cm = col_l2.number_input("Alto (cm):", min_value=1.0, value=9.0, step=0.5)
    L_lad, H_lad = l_cm / 100.0, h_cm / 100.0

# MÓDULO DE REVESTIMIENTOS Y CERÁMICOS (CON OPCIÓN PERSONALIZADA)
st.sidebar.write("---")
st.sidebar.subheader("📐 Revestimientos y Cerámicos")

tipo_cer_piso = st.sidebar.selectbox(
    "Formato Cerámica de Piso:",
    ["Cerámica de Piso (40x40 cm)", "Porcelanato de Piso (60x60 cm)", "Cerámica de Piso (30x30 cm)", "Otro (Personalizado)"]
)
if tipo_cer_piso == "Otro (Personalizado)":
    col_p1, col_p2 = st.sidebar.columns(2)
    piso_l = col_p1.number_input("Ancho Piso (cm):", min_value=1.0, value=45.0, step=1.0, key="custom_pl")
    piso_h = col_p2.number_input("Largo Piso (cm):", min_value=1.0, value=45.0, step=1.0, key="custom_ph")
    nombre_piso = f"Personalizado ({int(piso_l)}x{int(piso_h)} cm)"
else:
    nombre_piso = tipo_cer_piso

tipo_cer_pared = st.sidebar.selectbox(
    "Formato Cerámica de Pared:",
    ["Cerámica de Pared (20x30 cm)", "Cerámica de Pared (30x30 cm)", "Porcelanato de Pared (60x60 cm)", "Otro (Personalizado)"]
)
if tipo_cer_pared == "Otro (Personalizado)":
    col_w1, col_w2 = st.sidebar.columns(2)
    pared_l = col_w1.number_input("Ancho Pared (cm):", min_value=1.0, value=25.0, step=1.0, key="custom_wl")
    pared_h = col_w2.number_input("Largo Pared (cm):", min_value=1.0, value=40.0, step=1.0, key="custom_wh")
    nombre_pared = f"Personalizado ({int(pared_l)}x{int(pared_h)} cm)"
else:
    nombre_pared = tipo_cer_pared

# FACTORES LOGÍSTICOS NORMADOS CORREGIDOS
JUNTA_LADRILLO = 0.015       # 1.5 cm de junta
DESPERDECIO_LADRILLO = 0.05  # 5% merma muros
DESPERDECIO_TECHO = 0.05     # 5% merma ladrillo techo
FACTOR_CEMENTO_M2 = 2.5      # 2.5 bolsas por m²
FACTOR_ACERO_M2 = 10.0       # 10 kg de fierro por m²
PESO_VARILLA_PROMEDIO = 7.0  # Peso promedio de varilla de 9m
st.sidebar.write("---")

# CONFIGURACIÓN PRECIOS PROVEEDOR 1
st.sidebar.header("🏪 Precios - PROVEEDOR 1")
p1_lad_muro = st.sidebar.number_input("Ladrillo Muro P1 (S/. por und):", min_value=0.1, value=0.90, step=0.05, key="p1_lm")
p1_lad_techo = st.sidebar.number_input("Ladrillo Techo P1 (S/. por und):", min_value=0.1, value=2.80, step=0.10, key="p1_lt")
p1_cemento = st.sidebar.number_input("Cemento P1 (S/. por bolsa):", min_value=1.0, value=28.00, step=0.50, key="p1_cem")
p1_fierro = st.sidebar.number_input("Fierro Corrugado P1 (S/. por varilla):", min_value=1.0, value=31.00, step=1.00, key="p1_fiero")
p1_cer_piso = st.sidebar.number_input("Cerámica Piso P1 (S/. por m²):", min_value=1.0, value=35.00, step=1.00, key="p1_cp")
p1_cer_pared = st.sidebar.number_input("Cerámica Pared P1 (S/. por m²):", min_value=1.0, value=30.00, step=1.00, key="p1_cw")

st.sidebar.write("---")

# CONFIGURACIÓN PRECIOS PROVEEDOR 2
st.sidebar.header("🏪 Precios - PROVEEDOR 2")
p2_lad_muro = st.sidebar.number_input("Ladrillo Muro P2 (S/. por und):", min_value=0.1, value=0.85, step=0.05, key="p2_lm")
p2_lad_techo = st.sidebar.number_input("Ladrillo Techo P2 (S/. por und):", min_value=0.1, value=2.95, step=0.10, key="p2_lt")
p2_cemento = st.sidebar.number_input("Cemento P2 (S/. por bolsa):", min_value=1.0, value=29.50, step=0.50, key="p2_cem")
p2_fierro = st.sidebar.number_input("Fierro Corrugado P2 (S/. por varilla):", min_value=1.0, value=29.00, step=1.00, key="p2_fiero")
p2_cer_piso = st.sidebar.number_input("Cerámica Piso P2 (S/. por m²):", min_value=1.0, value=32.00, step=1.00, key="p2_cp")
p2_cer_pared = st.sidebar.number_input("Cerámica Pared P2 (S/. por m²):", min_value=1.0, value=34.00, step=1.00, key="p2_cw")


# 3. ALGORITMO CIENTÍFICO DE METRADOS MULTIPISO
# Uso directo de la variable calculada
area_total_acumulada = area_total_construir

# A. Ladrillos de Muro
cant_ladrillos_m2_neto = 1 / ((L_lad + JUNTA_LADRILLO) * (H_lad + JUNTA_LADRILLO))
total_ladrillos_muro = round((cant_ladrillos_m2_neto * area_total_acumulada) * (1 + DESPERDECIO_LADRILLO))

# B. Ladrillos de Techo (Aligerado)
total_ladrillos_techo = round((area_total_acumulada * 8.33) * (1 + DESPERDECIO_TECHO))

# C. Cemento
total_bolsas_cemento = round(area_total_acumulada * FACTOR_CEMENTO_M2, 1)

# D. Acero Estructural (Fierro en Varillas)
peso_acero_total_kg = area_total_acumulada * FACTOR_ACERO_M2
total_varillas_fierro = round(peso_acero_total_kg / PESO_VARILLA_PROMEDIO)

# E. Acabados (Cerámicos con mermas de 5% y 10%)
total_m2_piso = round(area_total_acumulada * 1.05, 1)
total_m2_pared = round(area_pared * 1.10, 1)


# 4. COSTOS TOTALES EN PARALELO
# Proveedor 1
c_lm_p1 = total_ladrillos_muro * p1_lad_muro
c_lt_p1 = total_ladrillos_techo * p1_lad_techo
c_cem_p1 = total_bolsas_cemento * p1_cemento
c_f_p1 = total_varillas_fierro * p1_fierro
c_cp_p1 = total_m2_piso * p1_cer_piso
c_cw_p1 = total_m2_pared * p1_cer_pared
total_p1 = c_lm_p1 + c_lt_p1 + c_cem_p1 + c_f_p1 + c_cp_p1 + c_cw_p1

# Proveedor 2
c_lm_p2 = total_ladrillos_muro * p2_lad_muro
c_lt_p2 = total_ladrillos_techo * p2_lad_techo
c_cem_p2 = total_bolsas_cemento * p2_cemento
c_f_p2 = total_varillas_fierro * p2_fierro
c_cp_p2 = total_m2_piso * p2_cer_piso
c_cw_p2 = total_m2_pared * p2_cer_pared
total_p2 = c_lm_p2 + c_lt_p2 + c_cem_p2 + c_f_p2 + c_cp_p2 + c_cw_p2


# 5. VISUALIZACIÓN DE SUSTENTO TÉCNICO
st.subheader("📚 Sustento y Ratios de Ingeniería Civil")
with st.expander("Ver criterios analíticos y fórmulas de metrado estructural (Normativa CAPECO)"):
    st.markdown("### 1. Cantidad de Ladrillos de Muro por m²")
    st.latex(r"C_{muro} = \frac{1}{(L_{lad} + J) \times (H_{lad} + J)} \times (1 + \%M_{muro}) \times N_{pisos}")
    st.markdown(f"- **Formato seleccionado:** {tipo_ladrillo}")
    st.markdown(f"- **Rendimiento base calculado:** {cant_ladrillos_m2_neto:.2f} und/m² (Con junta de {JUNTA_LADRILLO*100} cm y {DESPERDECIO_LADRILLO*100}% de merma).")
    
    st.markdown("### 2. Cantidad de Ladrillos de Techo Aligerado")
    st.latex(r"C_{techo} = \left( A_{piso} \times N_{pisos} \times 8.33 \text{ und/m}^2 \right) \times (1 + \%M_{techo})")
    
    st.markdown("### 3. Acero Estructural (Fierro Corrugado)")
    st.latex(r"V_{fierro} = \frac{(A_{piso} \times N_{pisos}) \times \text{Cuantía Estructural (10 kg/m}^2\text{)}}{\text{Peso Comercial de Varilla (7 kg/9m)}}")
    
    st.markdown("### 4. Revestimientos Cerámicos (Pisos y Paredes)")
    st.latex(r"\text{Metrado Piso} = (A_{piso} \times N_{pisos}) \times 1.05 \quad | \quad \text{Metrado Pared} = A_{enchape} \times 1.10")
    st.markdown(f"- **Piso:** {nombre_piso} (Añade 5% de merma por cortes multiplicada por los {num_pisos} niveles).")
    st.markdown(f"- **Pared:** {nombre_pared} (Añade 10% de merma debido a esquinas y derrames).")

st.write("---")

# 6. RESUMEN DE METRADOS
st.subheader(f"📊 Cantidades Totales Requeridas (Lista de Materiales - Edificación de {num_pisos} Pisos)")
col1, col2, col3 = st.columns(3)
col1.metric("🧱 Ladrillos de Muro Total", f"{total_ladrillos_muro} und")
col1.metric("🏠 Ladrillos de Techo Total", f"{total_ladrillos_techo} und")

col2.metric("🪨 Cemento Sol / APU Total", f"{total_bolsas_cemento} bolsas")
col2.metric("⛓️ Fierro de Construcción Total", f"{total_varillas_fierro} varillas (9m)")

col3.metric("📐 Cerámica de Piso Total", f"{total_m2_piso} m²")
col3.metric("🧼 Cerámica de Pared Total", f"{total_m2_pared} m²")

st.write("---")

# 7. ANÁLISIS COMPARATIVO DE PRESUPUESTOS
st.subheader("⚖️ Análisis Comparativo de Presupuestos")

es_p1_menor = total_p1 < total_p2
es_p2_menor = total_p2 < total_p1
diferencia_absoluta = abs(total_p1 - total_p2)

cm1, cm2 = st.columns(2)

with cm1:
    st.markdown("### 🏪 Proveedor 1")
    if es_p1_menor:
        st.metric(
            label="🏆 ¡OPCIÓN MÁS ECONÓMICA!", 
            value=f"S/. {total_p1:,.2f}",
            delta=f"- S/. {diferencia_absoluta:,.2f} (Ahorro óptimo)"
        )
    else:
        st.metric(
            label="COSTO TOTAL ESTIMADO P1", 
            value=f"S/. {total_p1:,.2f}",
            delta=f"+ S/. {diferencia_absoluta:,.2f} (Más caro)"
        )

with cm2:
    st.markdown("### 🏪 Proveedor 2")
    if es_p2_menor:
        st.metric(
            label="🏆 ¡OPCIÓN MÁS ECONÓMICA!", 
            value=f"S/. {total_p2:,.2f}",
            delta=f"- S/. {diferencia_absoluta:,.2f} (Ahorro óptimo)"
        )
    else:
        st.metric(
            label="COSTO TOTAL ESTIMADO P2", 
            value=f"S/. {total_p2:,.2f}",
            delta=f"+ S/. {diferencia_absoluta:,.2f} (Más caro)"
        )

st.write("##")

st.markdown("#### 📊 Desglose de Costos Estructurado")

df_comparativo = pd.DataFrame({
    "Material de Construcción": [
        "🧱 Ladrillo Muro", 
        "🏠 Ladrillo Techo", 
        "🪨 Cemento Sol / APU", 
        "⛓️ Fierro (Varillas)", 
        f"📐 Piso ({nombre_piso})", 
        f"🧼 Pared ({nombre_pared})"
    ],
    "Cantidad Requerida": [
        f"{total_ladrillos_muro} und", 
        f"{total_ladrillos_techo} und", 
        f"{total_bolsas_cemento} bolsas", 
        f"{total_varillas_fierro} var", 
        f"{total_m2_piso} m²", 
        f"{total_m2_pared} m²"
    ],
    "P1: P. Unitario": [f"S/. {p1_lad_muro:.2f}", f"S/. {p1_lad_techo:.2f}", f"S/. {p1_cemento:.2f}", f"S/. {p1_fierro:.2f}", f"S/. {p1_cer_piso:.2f}", f"S/. {p1_cer_pared:.2f}"],
    "P1: Subtotal": [f"S/. {c_lm_p1:,.2f}", f"S/. {c_lt_p1:,.2f}", f"S/. {c_cem_p1:,.2f}", f"S/. {c_f_p1:,.2f}", f"S/. {c_cp_p1:,.2f}", f"S/. {c_cw_p1:,.2f}"],
    "P2: P. Unitario": [f"S/. {p2_lad_muro:.2f}", f"S/. {p2_lad_techo:.2f}", f"S/. {p2_cemento:.2f}", f"S/. {p2_fierro:.2f}", f"S/. {p2_cer_piso:.2f}", f"S/. {p2_cer_pared:.2f}"],
    "P2: Subtotal": [f"S/. {c_lm_p2:,.2f}", f"S/. {c_lt_p2:,.2f}", f"S/. {c_cem_p2:,.2f}", f"S/. {c_f_p2:,.2f}", f"S/. {c_cp_p2:,.2f}", f"S/. {c_cw_p2:,.2f}"]
})

st.dataframe(df_comparativo, use_container_width=True, hide_index=True)

# 8. RECOMENDACIÓN DE COMPRA
st.subheader("💡 Recomendación de Optimización Financiera")
if total_p1 < total_p2:
    ahorro = total_p2 - total_p1
    st.success(f"✔️ **Estrategia de Suministro:** Se recomienda realizar la compra con el **Proveedor 1**. Ahorro neto de **S/. {ahorro:,.2f}**.")
elif total_p2 < total_p1:
    ahorro = total_p1 - total_p2
    st.success(f"✔️ **Estrategia de Suministro:** Se recomienda realizar la compra con el **Proveedor 2**. Ahorro neto de **S/. {ahorro:,.2f}**.")
else:
    st.info("📊 Ambos proveedores presentan un empate técnico.")

# =============================================================================
# GRÁFICO COMPARATIVO
# =============================================================================
st.subheader("📊 Comparativa Visual por Material")

df_grafico = pd.DataFrame({
    "Material": ["Ladrillo Muro", "Ladrillo Techo", "Cemento", "Fierro", "Piso", "Pared"],
    "Proveedor 1": [c_lm_p1, c_lt_p1, c_cem_p1, c_f_p1, c_cp_p1, c_cw_p1],
    "Proveedor 2": [c_lm_p2, c_lt_p2, c_cem_p2, c_f_p2, c_cp_p2, c_cw_p2]
})

df_melted = df_grafico.melt(id_vars="Material", var_name="Proveedor", value_name="Costo (S/.)")

fig = px.bar(df_melted, x="Material", y="Costo (S/.)", color="Proveedor",
             barmode="group", title="Costo por Material - P1 vs P2")
st.plotly_chart(fig, use_container_width=True)


# =============================================================================
# MÓDULO DE EXPORTACIÓN A EXCEL PREMIUM
# =============================================================================
st.markdown("---")
st.subheader("📥 Exportar Presupuesto")

def exportar_excel_profesIONAL(df_origen):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Presupuesto'
    
    ws.views.sheetView[0].showGridLines = True
    
    fuente_titulo = Font(name='Segoe UI', size=15, bold=True, color='FFFFFF')
    fuente_cabecera = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    fuente_datos = Font(name='Segoe UI', size=11, bold=False)
    fuente_total = Font(name='Segoe UI', size=11, bold=True, color='000000')
    
    fill_titulo = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    fill_cabecera = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    fill_total = PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid')
    
    borde_delgado = Side(border_style="thin", color="D0D3D4")
    border_celda = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
    border_doble = Border(bottom=Side(border_style="double", color="000000"), top=Side(border_style="thin", color="000000"))
    
    ws.merge_cells('A1:F2')
    ws['A1'] = "SISTEMA INTEGRAL DE PRESUPUESTOS - CAPECO"
    ws['A1'].font = fuente_titulo
    ws['A1'].fill = fill_titulo
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    columnas = ["Material de Construcción", "Cantidad Requerida", "P1: P. Unitario", "P1: Subtotal", "P2: P. Unitario", "P2: Subtotal"]
    for col_num, header in enumerate(columnas, 1):
        celda = ws.cell(row=4, column=col_num, value=header)
        celda.font = fuente_cabecera
        celda.fill = fill_cabecera
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    fila_inicio = 5
    for idx, row in df_origen.iterrows():
        fila_act = fila_inicio + idx
        
        c_a = ws.cell(row=fila_act, column=1, value=str(row['Material de Construcción']))
        c_a.alignment = Alignment(horizontal='left', vertical='center')
        
        c_b = ws.cell(row=fila_act, column=2, value=str(row['Cantidad Requerida']))
        c_b.alignment = Alignment(horizontal='right', vertical='center')
        
        valores_origen = [row['P1: P. Unitario'], row['P1: Subtotal'], row['P2: P. Unitario'], row['P2: Subtotal']]
        
        valores_num = []
        for val in valores_origen:
            if isinstance(val, str):
                # Limpieza robusta eliminando S/., comas, espacios y caracteres especiales residuales
                val_limpio = val.replace('S/.', '').replace('S/', '').replace(',', '').replace(' ', '').strip()
                valores_num.append(float(val_limpio) if val_limpio else 0.0)
            else:
                valores_num.append(float(val) if val is not None else 0.0)
                
        for i, val_final in enumerate(valores_num, 3):
            c_num = ws.cell(row=fila_act, column=i, value=val_final)
            c_num.number_format = '"S/." #,##0.00'
            c_num.alignment = Alignment(horizontal='right', vertical='center')
            
        for col_num in range(1, 7):
            ws.cell(row=fila_act, column=col_num).font = fuente_datos
            ws.cell(row=fila_act, column=col_num).border = border_celda

    fila_total = fila_inicio + len(df_origen)
    ws.cell(row=fila_total, column=1, value="COSTO TOTAL ESTIMADO").font = fuente_total
    ws.cell(row=fila_total, column=1).fill = fill_total
    
    for col_num in range(2, 7):
        ws.cell(row=fila_total, column=col_num).fill = fill_total
        
    tot_p1 = ws.cell(row=fila_total, column=4, value=f"=SUM(D5:D{fila_total-1})")
    tot_p1.number_format = '"S/." #,##0.00'
    tot_p1.font = fuente_total
    tot_p1.border = border_doble
    
    tot_p2 = ws.cell(row=fila_total, column=6, value=f"=SUM(F5:F{fila_total-1})")
    tot_p2.number_format = '"S/." #,##0.00'
    tot_p2.font = fuente_total
    tot_p2.border = border_doble

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: continue
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)
        
    wb.save(output)
    return output.getvalue()

excel_data = exportar_excel_profesIONAL(df_comparativo)

st.download_button(
    label="⬇️ Descargar Presupuesto en Excel",
    data=excel_data,
    file_name="presupuesto_estructurado_capeco.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
